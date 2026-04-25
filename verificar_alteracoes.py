"""Compara a extracao atual com o estado salvo da execucao anterior, detecta:
  - editais novos (nao existiam antes)
  - editais atualizados (PDF republicado E algum campo nosso mudou)
  - editais removidos (nao estao mais abertos na FAPES)

Para cada edital novo ou atualizado, envia um e-mail e marca a coluna
"Status" na planilha indicando que o aviso ja foi enviado.

Le as credenciais SMTP do .env (local) ou variaveis de ambiente (GitHub Actions).
Atualiza editais_fapes/_state.json e gera editais_fapes/_extracao.xlsx.
"""

from __future__ import annotations

import json
import os
import re
import smtplib
import sys
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from extrair_editais_gemini import (
    CATEGORIA_DISPLAY,
    acoes_resumidas,
    formatar_evento,
    proxima_acao,
    somente_acoes,
)

SUBMISSAO_RE = re.compile(r"(submiss[aã]o|envio.*proposta|abertura.*chamada)", re.IGNORECASE)
NAO_PROPOSTA_RE = re.compile(
    r"(d[uú]vida|recurso|questionamento|impugna[cç][aã]o|"
    r"complementa[cç][aã]o|atendimento|documento)",
    re.IGNORECASE,
)


def is_submissao_proposta(ev: dict[str, Any]) -> bool:
    name = ev.get("evento") or ""
    return bool(SUBMISSAO_RE.search(name)) and not bool(NAO_PROPOSTA_RE.search(name))


def submissoes(cronograma: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [ev for ev in (cronograma or []) if is_submissao_proposta(ev)]


def submissoes_resumidas(cronograma: list[dict[str, Any]]) -> str:
    return "\n".join(formatar_evento(ev) for ev in submissoes(cronograma))

ROOT = Path(__file__).resolve().parent
EDITAIS_DIR = ROOT / "editais_fapes"
EXTRACAO_JSON = EDITAIS_DIR / "_extracao.json"
STATE_PATH = EDITAIS_DIR / "_state.json"
EXTRACAO_XLSX = EDITAIS_DIR / "_extracao.xlsx"

TZ = ZoneInfo("America/Sao_Paulo")

CAMPOS_MONITORADOS = [
    ("objetivo", "Objetivo"),
    ("publico_alvo", "Público-alvo"),
    ("valor_total", "Valor total"),
    ("valor_por_proposta", "Valor por proposta"),
    ("contato", "Contato"),
]


def agora_iso() -> str:
    return datetime.now(TZ).isoformat(timespec="seconds")


def chave(reg: dict[str, Any]) -> str:
    return f"{reg['categoria']}||{reg['titulo']}"


def submissoes_por_datas(cronograma: list[dict[str, Any]]) -> dict[tuple, dict[str, Any]]:
    """Mapeia (data_inicio, data_fim) -> primeiro evento de submissao com essas datas.
    Comparar por datas evita ruido quando o Gemini reformula o nome do evento."""
    result: dict[tuple, dict[str, Any]] = {}
    for ev in submissoes(cronograma):
        key = (ev.get("data_inicio"), ev.get("data_fim"))
        if key not in result:
            result[key] = ev
    return result


def diff_extracao(antigo: dict[str, Any], novo: dict[str, Any]) -> list[str]:
    mudancas: list[str] = []
    for campo, label in CAMPOS_MONITORADOS:
        v_antigo = antigo.get(campo)
        v_novo = novo.get(campo)
        if v_antigo != v_novo:
            mudancas.append(f"{label}: \"{v_antigo or '-'}\" → \"{v_novo or '-'}\"")

    old = submissoes_por_datas(antigo.get("cronograma") or [])
    new = submissoes_por_datas(novo.get("cronograma") or [])
    for key in sorted(new.keys() - old.keys(), key=lambda k: (k[0] or "", k[1] or "")):
        ev = new[key]
        mudancas.append(
            f"Submissão (nova): {ev.get('evento', '')} — {key[0] or '?'} a {key[1] or '?'}"
        )
    for key in sorted(old.keys() - new.keys(), key=lambda k: (k[0] or "", k[1] or "")):
        ev = old[key]
        mudancas.append(
            f"Submissão (removida): {ev.get('evento', '')} — {key[0] or '?'} a {key[1] or '?'}"
        )
    return mudancas


def status_label(state_entry: dict[str, Any]) -> str:
    s = state_entry.get("status_atual", "")
    when = state_entry.get("email_enviado_em")
    if when:
        try:
            dt = datetime.fromisoformat(when).strftime("%d/%m/%Y")
        except Exception:
            dt = when[:10]
    else:
        dt = None

    if s in ("novo", "atualizado") and dt is None:
        return "Novo - aguardando envio"
    if s == "novo":
        return f"Novo - e-mail enviado em {dt}"
    if s == "atualizado":
        return f"Atualizado - e-mail enviado em {dt}"
    if s == "removido":
        return "Removido (nao mais aberto)"
    return "Sem alteracoes"


def carregar_state() -> dict[str, Any]:
    if STATE_PATH.exists():
        with open(STATE_PATH, encoding="utf-8") as fh:
            return json.load(fh)
    return {"editais": {}, "ultima_execucao": None}


def salvar_state(state: dict[str, Any]) -> None:
    state["ultima_execucao"] = agora_iso()
    with open(STATE_PATH, "w", encoding="utf-8") as fh:
        json.dump(state, fh, ensure_ascii=False, indent=2)


def montar_email_novo(reg: dict[str, Any]) -> tuple[str, str, str]:
    cat = CATEGORIA_DISPLAY.get(reg["categoria"], reg["categoria"])
    titulo = reg["titulo"]
    ext = reg.get("extracao") or {}
    cron = ext.get("cronograma") or []
    subs = submissoes(cron)

    subject = f"[FAPES] Novo edital: {titulo}"

    subs_plain = "\n".join(f"  - {formatar_evento(ev)}" for ev in subs) or "  (sem datas de submissao identificadas)"
    plain = f"""Categoria: {cat}
Edital: {titulo}

OBJETIVO
{ext.get('objetivo') or '-'}

PUBLICO-ALVO: {ext.get('publico_alvo') or '-'}
VALOR TOTAL: {ext.get('valor_total') or '-'}
VALOR POR PROPOSTA: {ext.get('valor_por_proposta') or '-'}
CONTATO: {ext.get('contato') or '-'}

DATAS DE SUBMISSAO DE PROPOSTAS
{subs_plain}

PDF: {reg.get('pdf_url', '')}
"""

    subs_html = "".join(f"<li>{formatar_evento(ev)}</li>" for ev in subs) \
        or "<li><em>(sem datas de submissão identificadas)</em></li>"
    html = f"""<!DOCTYPE html><html><body style="font-family:sans-serif;max-width:680px">
<h2 style="color:#1a4d7a">Novo edital FAPES</h2>
<p><strong>Categoria:</strong> {cat}<br>
<strong>Edital:</strong> {titulo}</p>

<h3>Objetivo</h3>
<p>{(ext.get('objetivo') or '-').replace(chr(10), '<br>')}</p>

<table style="border-collapse:collapse">
<tr><td><strong>Público-alvo:</strong></td><td>{ext.get('publico_alvo') or '-'}</td></tr>
<tr><td><strong>Valor total:</strong></td><td>{ext.get('valor_total') or '-'}</td></tr>
<tr><td><strong>Valor por proposta:</strong></td><td>{ext.get('valor_por_proposta') or '-'}</td></tr>
<tr><td><strong>Contato:</strong></td><td>{ext.get('contato') or '-'}</td></tr>
</table>

<h3>Datas de submissão de propostas</h3>
<ul>{subs_html}</ul>

<p><a href="{reg.get('pdf_url', '')}">Baixar PDF do edital</a></p>
</body></html>"""

    return subject, plain, html


def montar_email_atualizado(
    reg: dict[str, Any],
    mudancas: list[str],
    novas_alteracoes: list[dict[str, str]],
) -> tuple[str, str, str]:
    cat = CATEGORIA_DISPLAY.get(reg["categoria"], reg["categoria"])
    titulo = reg["titulo"]

    subject = f"[FAPES] Edital atualizado: {titulo}"

    blocos_plain: list[str] = [f"Categoria: {cat}", f"Edital: {titulo}", ""]
    if novas_alteracoes:
        blocos_plain.append("DOCUMENTO(S) DE ALTERACAO PUBLICADO(S):")
        for a in novas_alteracoes:
            blocos_plain.append(f"  - {a.get('titulo', '(sem titulo)')}")
            if a.get("url"):
                blocos_plain.append(f"    {a['url']}")
        blocos_plain.append("")

    blocos_plain.append("ALTERACOES DETECTADAS NOS CAMPOS MONITORADOS:")
    if mudancas:
        for m in mudancas:
            blocos_plain.append(f"  - {m}")
    else:
        blocos_plain.append("  (nenhuma alteracao em objetivo, publico-alvo, valores, contato ou datas de submissao)")
    blocos_plain.append("")
    blocos_plain.append(f"PDF original: {reg.get('pdf_url', '')}")
    plain = "\n".join(blocos_plain)

    html_alts = ""
    if novas_alteracoes:
        items = "".join(
            f'<li>{a.get("titulo", "(sem título)")}'
            + (f' — <a href="{a["url"]}">PDF</a>' if a.get("url") else "")
            + "</li>"
            for a in novas_alteracoes
        )
        html_alts = f"<h3>Documento(s) de alteração publicado(s)</h3><ul>{items}</ul>"

    if mudancas:
        html_changes = "".join(f"<li>{m}</li>" for m in mudancas)
    else:
        html_changes = "<li><em>(nenhuma alteração em objetivo, público-alvo, valores, contato ou datas de submissão)</em></li>"

    html = f"""<!DOCTYPE html><html><body style="font-family:sans-serif;max-width:680px">
<h2 style="color:#a0522d">Edital FAPES atualizado</h2>
<p><strong>Categoria:</strong> {cat}<br>
<strong>Edital:</strong> {titulo}</p>

{html_alts}

<h3>Alterações detectadas nos campos monitorados</h3>
<ul>{html_changes}</ul>

<p><a href="{reg.get('pdf_url', '')}">PDF original</a></p>
</body></html>"""

    return subject, plain, html


def enviar_email(subject: str, plain: str, html: str, smtp_cfg: dict[str, str]) -> None:
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = smtp_cfg["from"]
    msg["To"] = smtp_cfg["to"]
    msg.set_content(plain)
    msg.add_alternative(html, subtype="html")

    with smtplib.SMTP(smtp_cfg["host"], int(smtp_cfg["port"])) as srv:
        srv.starttls()
        srv.login(smtp_cfg["user"], smtp_cfg["password"])
        srv.send_message(msg)


def carregar_smtp() -> dict[str, str] | None:
    user = os.getenv("SMTP_USER")
    pw = os.getenv("SMTP_PASSWORD")
    to = os.getenv("EMAIL_TO")
    if not (user and pw and to):
        return None
    return {
        "host": os.getenv("SMTP_HOST", "smtp.gmail.com"),
        "port": os.getenv("SMTP_PORT", "587"),
        "user": user,
        "password": pw,
        "from": user,
        "to": to,
    }


def gerar_xlsx(consolidados: list[dict[str, Any]], state: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Editais"
    headers = [
        "Categoria",
        "Edital",
        "Status",
        "Próxima ação do proponente",
        "Alterações detectadas",
        "Objetivo",
        "Público-alvo",
        "Valor total",
        "Valor por proposta",
        "Contato",
        "Ações do proponente (todas)",
        "Observações",
        "PDF (URL)",
        "Arquivo local",
    ]
    ws.append(headers)

    fill_novo = PatternFill("solid", fgColor="C6EFCE")
    fill_atualizado = PatternFill("solid", fgColor="FFEB9C")

    for reg in consolidados:
        ext = reg.get("extracao") or {}
        cron = ext.get("cronograma") or []
        ent = state["editais"].get(chave(reg), {})
        status_txt = status_label(ent)
        alteracoes = "\n".join(ent.get("alteracoes_recentes") or [])
        ws.append([
            CATEGORIA_DISPLAY.get(reg.get("categoria", ""), reg.get("categoria", "")),
            reg.get("titulo", ""),
            status_txt,
            proxima_acao(cron),
            alteracoes,
            ext.get("objetivo", ""),
            ext.get("publico_alvo", ""),
            ext.get("valor_total", ""),
            ext.get("valor_por_proposta", ""),
            ext.get("contato", ""),
            acoes_resumidas(cron),
            ext.get("observacoes_gerais", ""),
            reg.get("pdf_url", ""),
            reg.get("arquivo_local", ""),
        ])
        if status_txt.startswith("Novo"):
            ws.cell(row=ws.max_row, column=3).fill = fill_novo
        elif status_txt.startswith("Atualizado"):
            ws.cell(row=ws.max_row, column=3).fill = fill_atualizado

    widths = {"A": 22, "B": 55, "C": 38, "D": 50, "E": 45,
              "F": 60, "G": 35, "H": 18, "I": 18, "J": 35,
              "K": 60, "L": 40, "M": 55, "N": 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    bold = Font(bold=True)
    head_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    body_align = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = head_align
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_align
    ws.freeze_panes = "C2"
    wb.save(EXTRACAO_XLSX)


def main() -> None:
    load_dotenv(ROOT / ".env")

    if not EXTRACAO_JSON.exists():
        sys.exit(f"Nao encontrei {EXTRACAO_JSON}. Rode antes: python3 extrair_editais_gemini.py")

    with open(EXTRACAO_JSON, encoding="utf-8") as fh:
        consolidados = json.load(fh)

    state = carregar_state()
    smtp = carregar_smtp()
    if smtp is None:
        print("[aviso] Variaveis SMTP_* / EMAIL_TO ausentes - e-mails NAO serao enviados.")

    chaves_atuais = set()
    reg_por_chave: dict[str, dict] = {}

    for reg in consolidados:
        k = chave(reg)
        chaves_atuais.add(k)
        reg_por_chave[k] = reg
        ext_novo = reg.get("extracao") or {}
        alts_atual = reg.get("alteracoes") or []
        urls_atuais = [a["url"] for a in alts_atual if a.get("url")]
        ent = state["editais"].get(k)

        if ent is None:
            state["editais"][k] = {
                "categoria": reg["categoria"],
                "titulo": reg["titulo"],
                "pdf_url": reg.get("pdf_url"),
                "pdf_sha256": reg.get("pdf_sha256"),
                "alteracoes": [
                    {"url": a.get("url"), "titulo": a.get("titulo"),
                     "primeira_visualizacao": agora_iso()}
                    for a in alts_atual
                ],
                "primeira_visualizacao": agora_iso(),
                "ultima_visualizacao": agora_iso(),
                "status_atual": "novo",
                "email_enviado_em": None,
                "ultima_alteracao_em": agora_iso(),
                "alteracoes_recentes": [],
                "novas_alteracoes_doc": [],
                "extracao": ext_novo,
            }
            continue

        ent["ultima_visualizacao"] = agora_iso()
        ent["pdf_url"] = reg.get("pdf_url")
        ent["pdf_sha256"] = reg.get("pdf_sha256")

        urls_state = [a["url"] for a in (ent.get("alteracoes") or []) if a.get("url")]
        novas_urls = [u for u in urls_atuais if u not in urls_state]
        novos_docs = [
            {"url": a["url"], "titulo": a.get("titulo", "")}
            for a in alts_atual if a.get("url") in novas_urls
        ]

        ent["alteracoes"] = [
            {
                "url": a.get("url"),
                "titulo": a.get("titulo"),
                "primeira_visualizacao": next(
                    (s.get("primeira_visualizacao") for s in (ent.get("alteracoes") or [])
                     if s.get("url") == a.get("url")),
                    agora_iso(),
                ),
            }
            for a in alts_atual
        ]

        mudancas_campos = diff_extracao(ent.get("extracao") or {}, ext_novo)
        ent["extracao"] = ext_novo

        ja_emailed = bool(ent.get("email_enviado_em"))

        if mudancas_campos or novos_docs:
            if not ja_emailed:
                ent["status_atual"] = "novo"
                ent["alteracoes_recentes"] = []
                ent["novas_alteracoes_doc"] = []
            else:
                ent["status_atual"] = "atualizado"
                ent["email_enviado_em"] = None
                ent["alteracoes_recentes"] = mudancas_campos
                ent["novas_alteracoes_doc"] = novos_docs
            ent["ultima_alteracao_em"] = agora_iso()
        elif ent.get("status_atual") == "novo" and ja_emailed:
            ent["status_atual"] = "sem_alteracoes"
            ent["alteracoes_recentes"] = []
            ent["novas_alteracoes_doc"] = []

    for k, ent in state["editais"].items():
        if k not in chaves_atuais and ent.get("status_atual") != "removido":
            ent["status_atual"] = "removido"
            ent["alteracoes_recentes"] = []
            ent["novas_alteracoes_doc"] = []

    pendentes_novos: list[tuple[dict, str]] = []
    pendentes_atualizados: list[tuple[dict, str, list[str], list[dict]]] = []
    for k in chaves_atuais:
        ent = state["editais"][k]
        if ent.get("status_atual") in ("sem_alteracoes", "removido"):
            continue
        reg = reg_por_chave[k]
        ja_emailed = bool(ent.get("email_enviado_em"))
        if not ja_emailed:
            pendentes_novos.append((reg, k))
        elif ent.get("status_atual") == "atualizado":
            pendentes_atualizados.append((
                reg, k,
                ent.get("alteracoes_recentes") or [],
                ent.get("novas_alteracoes_doc") or [],
            ))

    print(f"\n  Novos pendentes:       {len(pendentes_novos)}")
    print(f"  Atualizados pendentes: {len(pendentes_atualizados)}")
    print(f"  Total atual:           {len(consolidados)}")

    if smtp:
        for reg, k in pendentes_novos:
            subject, plain, html = montar_email_novo(reg)
            try:
                enviar_email(subject, plain, html, smtp)
                state["editais"][k]["email_enviado_em"] = agora_iso()
                print(f"  [email novo] {reg['titulo'][:60]}")
            except Exception as exc:
                print(f"  [erro email novo] {reg['titulo'][:60]}: {exc}")

        for reg, k, mudancas, novos_docs in pendentes_atualizados:
            subject, plain, html = montar_email_atualizado(reg, mudancas, novos_docs)
            try:
                enviar_email(subject, plain, html, smtp)
                state["editais"][k]["email_enviado_em"] = agora_iso()
                print(f"  [email atualizado] {reg['titulo'][:60]}")
            except Exception as exc:
                print(f"  [erro email atual] {reg['titulo'][:60]}: {exc}")

    salvar_state(state)
    gerar_xlsx(consolidados, state)
    print(f"\nState atualizado: {STATE_PATH}")
    print(f"Planilha:         {EXTRACAO_XLSX}")


if __name__ == "__main__":
    main()
